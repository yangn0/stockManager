from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, session
from functools import wraps
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from urllib.parse import quote
import database
from secret import USERNAME,PASSWORD

app = Flask(__name__)
app.secret_key = 'stock_manager_secret_key'

CATEGORIES = ['耐克衣服', '耐克鞋子', '耐克配件', '阿迪衣服', '阿迪鞋子', '阿迪配件', '李宁衣服', '李宁鞋子', '李宁配件']


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username == USERNAME and password == PASSWORD:
            session['logged_in'] = True
            flash('登录成功！', 'success')
            return redirect(url_for('index'))
        else:
            flash('用户名或密码错误！', 'error')

    return render_template('login.html')


@app.route('/logout')
def logout():
    """登出"""
    session.pop('logged_in', None)
    flash('已退出登录！', 'success')
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    """首页 - 库存展示"""
    category = request.args.get('category', 'all')
    inventory = database.get_inventory(category)
    total_value = database.get_total_value()
    return render_template('index.html',
                           inventory=inventory,
                           total_value=total_value,
                           categories=CATEGORIES,
                           selected_category=category)


@app.route('/stock_in', methods=['GET', 'POST'])
@login_required
def stock_in():
    """入库页面"""
    if request.method == 'POST':
        category = request.form.get('category')
        product_code = request.form.get('product_code')
        size = request.form.get('size')
        purchase_price = float(request.form.get('purchase_price'))
        quantity = int(request.form.get('quantity'))

        database.add_stock(category, product_code, size, purchase_price, quantity)
        flash('入库成功！', 'success')
        return redirect(url_for('stock_in'))

    return render_template('stock_in.html', categories=CATEGORIES)


@app.route('/stock_out', methods=['GET', 'POST'])
@login_required
def stock_out():
    """出库页面"""
    search_results = []
    search_code = request.args.get('search', '')

    if search_code:
        search_results = database.search_by_product_code(search_code)

    return render_template('stock_out.html',
                           search_results=search_results,
                           search_code=search_code)


@app.route('/do_stock_out', methods=['POST'])
@login_required
def do_stock_out():
    """执行出库操作"""
    item_id = int(request.form.get('item_id'))
    sell_price = float(request.form.get('sell_price'))
    quantity = int(request.form.get('quantity'))

    success, message = database.remove_stock(item_id, sell_price, quantity)

    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')

    return redirect(url_for('stock_out'))


@app.route('/delete_inventory', methods=['POST'])
@login_required
def delete_inventory():
    """删除库存"""
    item_id = int(request.form.get('item_id'))
    quantity = int(request.form.get('quantity'))
    category = request.form.get('category', 'all')

    success, message = database.delete_inventory(item_id, quantity)

    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')

    return redirect(url_for('index', category=category))


@app.route('/records')
@login_required
def records():
    """出入库记录"""
    stock_in_records = database.get_stock_in_records()
    stock_out_records = database.get_stock_out_records()
    return render_template('records.html',
                           stock_in_records=stock_in_records,
                           stock_out_records=stock_out_records)


@app.route('/monthly')
@login_required
def monthly():
    """月度汇总"""
    summary = database.get_monthly_summary()
    return render_template('monthly.html', summary=summary)


@app.route('/api/item/<int:item_id>')
@login_required
def get_item(item_id):
    """获取库存项详情API"""
    item = database.get_inventory_item(item_id)
    if item:
        return jsonify({
            'id': item['id'],
            'category': item['category'],
            'product_code': item['product_code'],
            'size': item['size'],
            'purchase_price': item['purchase_price'],
            'quantity': item['quantity']
        })
    return jsonify({'error': '未找到'}), 404


@app.route('/yearly')
@login_required
def yearly():
    """年度汇总"""
    summary = database.get_yearly_summary()
    return render_template('yearly.html', summary=summary)


def create_excel_style():
    """创建Excel样式"""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_font, header_fill, header_alignment, thin_border


@app.route('/export/inventory')
@login_required
def export_inventory():
    """导出当前库存为Excel"""
    category = request.args.get('category', 'all')
    inventory = database.get_inventory(category)
    total_value = database.get_total_value()

    wb = Workbook()
    ws = wb.active
    ws.title = "当前库存"

    header_font, header_fill, header_alignment, thin_border = create_excel_style()

    # 标题行
    headers = ['类别', '货号', '尺码', '进货价', '数量', '货值']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    for row_idx, item in enumerate(inventory, 2):
        ws.cell(row=row_idx, column=1, value=item['category']).border = thin_border
        ws.cell(row=row_idx, column=2, value=item['product_code']).border = thin_border
        ws.cell(row=row_idx, column=3, value=item['size']).border = thin_border
        ws.cell(row=row_idx, column=4, value=item['purchase_price']).border = thin_border
        ws.cell(row=row_idx, column=5, value=item['quantity']).border = thin_border
        ws.cell(row=row_idx, column=6, value=item['purchase_price'] * item['quantity']).border = thin_border

    # 合计行
    last_row = len(inventory) + 2
    ws.cell(row=last_row, column=5, value='总货值:').font = Font(bold=True)
    ws.cell(row=last_row, column=6, value=total_value).font = Font(bold=True)

    # 调整列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"库存_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.route('/export/stock_in')
@login_required
def export_stock_in():
    """导出入库记录为Excel"""
    records = database.get_stock_in_records()

    wb = Workbook()
    ws = wb.active
    ws.title = "入库记录"

    header_font, header_fill, header_alignment, thin_border = create_excel_style()

    headers = ['时间', '类别', '货号', '尺码', '进货价', '数量', '金额']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record['created_at']).border = thin_border
        ws.cell(row=row_idx, column=2, value=record['category']).border = thin_border
        ws.cell(row=row_idx, column=3, value=record['product_code']).border = thin_border
        ws.cell(row=row_idx, column=4, value=record['size']).border = thin_border
        ws.cell(row=row_idx, column=5, value=record['purchase_price']).border = thin_border
        ws.cell(row=row_idx, column=6, value=record['quantity']).border = thin_border
        ws.cell(row=row_idx, column=7, value=record['purchase_price'] * record['quantity']).border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"入库记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.route('/export/stock_out')
@login_required
def export_stock_out():
    """导出出库记录为Excel"""
    records = database.get_stock_out_records()

    wb = Workbook()
    ws = wb.active
    ws.title = "出库记录"

    header_font, header_fill, header_alignment, thin_border = create_excel_style()

    headers = ['时间', '类别', '货号', '尺码', '进货价', '卖出价', '数量', '利润']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record['created_at']).border = thin_border
        ws.cell(row=row_idx, column=2, value=record['category']).border = thin_border
        ws.cell(row=row_idx, column=3, value=record['product_code']).border = thin_border
        ws.cell(row=row_idx, column=4, value=record['size']).border = thin_border
        ws.cell(row=row_idx, column=5, value=record['purchase_price']).border = thin_border
        ws.cell(row=row_idx, column=6, value=record['sell_price']).border = thin_border
        ws.cell(row=row_idx, column=7, value=record['quantity']).border = thin_border
        profit_cell = ws.cell(row=row_idx, column=8, value=record['profit'])
        profit_cell.border = thin_border
        if record['profit'] < 0:
            profit_cell.font = Font(color='FF0000')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"出库记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.route('/export/monthly')
@login_required
def export_monthly():
    """导出月度汇总为Excel"""
    summary = database.get_monthly_summary()

    wb = Workbook()
    ws = wb.active
    ws.title = "月度汇总"

    header_font, header_fill, header_alignment, thin_border = create_excel_style()

    headers = ['月份', '销售额', '成本', '利润', '销量', '利润率']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    total_revenue = 0
    total_cost = 0
    total_profit = 0
    total_quantity = 0

    for row_idx, record in enumerate(summary, 2):
        ws.cell(row=row_idx, column=1, value=record['month']).border = thin_border
        ws.cell(row=row_idx, column=2, value=record['revenue']).border = thin_border
        ws.cell(row=row_idx, column=3, value=record['cost']).border = thin_border
        profit_cell = ws.cell(row=row_idx, column=4, value=record['total_profit'])
        profit_cell.border = thin_border
        if record['total_profit'] < 0:
            profit_cell.font = Font(color='FF0000')
        ws.cell(row=row_idx, column=5, value=record['total_quantity']).border = thin_border
        profit_rate = (record['total_profit'] / record['cost'] * 100) if record['cost'] else 0
        ws.cell(row=row_idx, column=6, value=f"{profit_rate:.1f}%").border = thin_border

        total_revenue += record['revenue'] or 0
        total_cost += record['cost'] or 0
        total_profit += record['total_profit'] or 0
        total_quantity += record['total_quantity'] or 0

    # 合计行
    last_row = len(summary) + 2
    ws.cell(row=last_row, column=1, value='合计').font = Font(bold=True)
    ws.cell(row=last_row, column=2, value=total_revenue).font = Font(bold=True)
    ws.cell(row=last_row, column=3, value=total_cost).font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total_profit).font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=total_quantity).font = Font(bold=True)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"月度汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.route('/export/yearly')
@login_required
def export_yearly():
    """导出年度汇总为Excel"""
    summary = database.get_yearly_summary()

    wb = Workbook()
    ws = wb.active
    ws.title = "年度汇总"

    header_font, header_fill, header_alignment, thin_border = create_excel_style()

    headers = ['年份', '销售额', '成本', '利润', '销量', '利润率']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    total_revenue = 0
    total_cost = 0
    total_profit = 0
    total_quantity = 0

    for row_idx, record in enumerate(summary, 2):
        ws.cell(row=row_idx, column=1, value=record['year']).border = thin_border
        ws.cell(row=row_idx, column=2, value=record['revenue']).border = thin_border
        ws.cell(row=row_idx, column=3, value=record['cost']).border = thin_border
        profit_cell = ws.cell(row=row_idx, column=4, value=record['total_profit'])
        profit_cell.border = thin_border
        if record['total_profit'] < 0:
            profit_cell.font = Font(color='FF0000')
        ws.cell(row=row_idx, column=5, value=record['total_quantity']).border = thin_border
        profit_rate = (record['total_profit'] / record['cost'] * 100) if record['cost'] else 0
        ws.cell(row=row_idx, column=6, value=f"{profit_rate:.1f}%").border = thin_border

        total_revenue += record['revenue'] or 0
        total_cost += record['cost'] or 0
        total_profit += record['total_profit'] or 0
        total_quantity += record['total_quantity'] or 0

    # 合计行
    last_row = len(summary) + 2
    ws.cell(row=last_row, column=1, value='合计').font = Font(bold=True)
    ws.cell(row=last_row, column=2, value=total_revenue).font = Font(bold=True)
    ws.cell(row=last_row, column=3, value=total_cost).font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total_profit).font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=total_quantity).font = Font(bold=True)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"年度汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


if __name__ == '__main__':
    database.init_db()
    app.run("0.0.0.0",debug=True, port=15000)
